from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ..models import ScrapingJob
import io
import re

class ExcelExporter:
    PROFILE_PATTERNS = [
        r'^/masterclass-profile/',
        r'^/speaker-profile/',
        r'^/adviser-profile/',
        r'^/masterclasses/',
        r'^/conferenciers/',
        r'^/orador/',
        r'^/speakers?/',
        r'^/profiles?/',
        r'^/experts?/',
        r'^/consultants?/',
        r'^/advisers?/',
        r'^/trainers?/',
        r'^/instructors?/'
    ]

    def __init__(self, job_id):
        self.job = ScrapingJob.objects.get(id=job_id)
        self.wb = Workbook()

    def _create_header_style(self):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        return header_font, header_fill

    def _is_profile_path(self, path):
        """Check if the path matches any of the profile patterns"""
        return any(re.match(pattern, path, re.IGNORECASE) for pattern in self.PROFILE_PATTERNS)

    def _add_profile_endpoints_sheet(self):
        """Add a sheet specifically for profile and speaker-related endpoints"""
        ws = self.wb.create_sheet("Profile Endpoints")
        
        # Add headers
        header_font, header_fill = self._create_header_style()
        headers = ["Endpoint Name", "URL", "Path", "Type", "Discovered At"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Filter and add profile-related endpoints
        row = 2
        for endpoint in self.job.endpoints.all():
            if self._is_profile_path(endpoint.path):
                # Determine the type based on the path
                endpoint_type = next(
                    (pattern.strip('^/').strip('/') for pattern in self.PROFILE_PATTERNS 
                     if re.match(pattern, endpoint.path, re.IGNORECASE)),
                    'other'
                )
                
                ws.cell(row=row, column=1, value=endpoint.endpoint_name)
                ws.cell(row=row, column=2, value=endpoint.url)
                ws.cell(row=row, column=3, value=endpoint.path)
                ws.cell(row=row, column=4, value=endpoint_type)
                ws.cell(row=row, column=5, value=endpoint.discovered_at.strftime('%Y-%m-%d %H:%M:%S'))
                row += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    def _add_endpoints_sheet(self):
        ws = self.wb.active
        ws.title = "Website Endpoints"
        
        # Add headers
        header_font, header_fill = self._create_header_style()
        headers = ["Endpoint Name", "URL", "Path", "Discovered At"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Add data
        for row, endpoint in enumerate(self.job.endpoints.all(), 2):
            ws.cell(row=row, column=1, value=endpoint.endpoint_name)
            ws.cell(row=row, column=2, value=endpoint.url)
            ws.cell(row=row, column=3, value=endpoint.path)
            ws.cell(row=row, column=4, value=endpoint.discovered_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    def _add_contents_sheet(self):
        ws = self.wb.create_sheet("Scraped Contents")
        
        # Add headers
        header_font, header_fill = self._create_header_style()
        headers = ["URL", "Scraped At", "File Path"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Add data
        for row, content in enumerate(self.job.contents.all(), 2):
            ws.cell(row=row, column=1, value=content.url)
            ws.cell(row=row, column=2, value=content.scraped_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=3, value=content.html_file_path)

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    def export(self):
        self._add_endpoints_sheet()
        self._add_profile_endpoints_sheet()  # Add the new profile endpoints sheet
        self._add_contents_sheet()
        
        # Save to buffer
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        return buffer
